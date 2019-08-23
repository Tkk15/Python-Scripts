import openpyxl
wb = openpyxl.load_workbook('CIS3100.01.xlsx')


wb.create_sheet('Groups')
s1=wb['Sheet1']
#s1.title ='Student Names'
s2=wb['Groups']

print(wb.sheetnames)




for i in range(2, s1.max_row, 1):
    counter = (i-1)%3
    print(i, s1.cell(row=i, column=2).value)  
    if(counter==0):
        s2.cell(row=i-3, column=2).value = str( 
          s1.cell(row=i,column=1).value + ' ' +     #First Name
          s1.cell(row=i,column=2).value + ', ' +    #Last Name
          s1.cell(row=i-1,column=1).value + ' ' + 
          s1.cell(row=i-1,column=2).value + ', ' +
          s1.cell(row=i-2,column=1).value  + ' ' +
          s1.cell(row=i-2,column=2).value )
        print(s2.cell(row=i-3, column=2).value)

wb.remove(wb['Sheet1'])
wb.save('Group123s.xlsx')



    
    
